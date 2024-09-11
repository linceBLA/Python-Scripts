# script to extract user login authentification cookies
# This script run through a list of URL, test multiple login credentials and identify the credentials working
# It then extracts the authentification cookies of the working credentials by username 
# As results a folder user_cookies will be created and contains the txt files of cookies for the credentials working for the pages.
# This script needs:
# - list of URL in excel table: urls_auth.xlsx
# - list of credentials pair in excel table: credentials.xlsx 
# - the ID names of input extracted from the URL in excel table. These ID names can be extarcted via another python script.
#	
# NOTA: If 1 unique username/password credentials is working for all urls, it won't collect the other credentials cookies 
# header: 'User-Agent': 'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.0)-HACKERNAME'



import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import os
from tqdm import tqdm

# Load the workbook with URLs
try:
    workbook = openpyxl.load_workbook("urls_auth.xlsx")
    url_sheet = workbook["Sheet1"]
except FileNotFoundError:
    print("File 'urls_auth.xlsx' not found.")
    exit(1)

# Load the workbook with credentials
try:
    credentials_workbook = openpyxl.load_workbook("credentials.xlsx")
    credentials_sheet = credentials_workbook["Sheet1"]
    credentials = [
        (cell.value, cell.offset(column=1).value)
        for cell in credentials_sheet["A"][1:]
        if cell.value is not None and cell.offset(column=1).value is not None
    ]
except FileNotFoundError:
    print("File 'credentials.xlsx' not found.")
    exit(1)

# Read input names IDs from an Excel file
df_ids = pd.read_excel('output-inputs.xlsx', sheet_name='input_ids', usecols=[0, 1], names=['username_TAG', 'password_TAG']) 
# NOTA: output-inputs.xlsx is the excel table of the list of ID names of input extracted from the URL
input_usernames = df_ids['username_TAG'].tolist()
input_passwords = df_ids['password_TAG'].tolist()

# Create a WebDriver instance
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)

headers = {'User-Agent': 'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.0)-HACKERNAME'}

# Define a function to check if a page is a login page
def is_login_page(html_content):
    keywords = ["login", "sign in", "username", "password"]
    for keyword in keywords:
        if keyword in html_content.lower():
            return True
    return False

# Define a function to check for a successful login based on HTTP status codes with headers
def check_login_status(url, headers, username, password):
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return "yes", ""
        elif response.status_code == 302:
            return "no", "Redirect to login page"
        else:
            return "no", "Unexpected status code"
    except requests.RequestException as e:
        return "no", str(e)

# Define a function to check if the user is already logged in
def is_user_already_logged_in(url, username):
    driver.get(url)
    page_source = driver.page_source
    return username in page_source

# Define a function to attempt login with a timeout handling
def attempt_login_with_timeout(url, username, password, timeout=30, headers=None):
    try:
        driver.set_page_load_timeout(timeout)
        if is_user_already_logged_in(url, username):
            return "yes", ""
        login_status, error = attempt_login(url, username, password, headers=headers)
        return login_status, error
    except TimeoutException:
        return "no", "Page load timeout"

# Define a function to attempt login
def attempt_login(url, username, password, headers=None):
    login_status, error = check_login_status(url, headers, username, password)

    if login_status == "no":
        return login_status, error

    driver.get(url)
    page_source = driver.page_source

    if is_user_already_logged_in(url, username):
        return "yes", ""
    elif is_login_page(page_source):
        # Find username and password input fields based on their names
        username_input = None
        password_input = None
        for input_name in input_usernames:
            try:
                username_input = driver.find_element(By.NAME, input_name)
                break
            except:
                continue
        for input_name in input_passwords:
            try:
                password_input = driver.find_element(By.NAME, input_name)
                break
            except:
                continue

        if username_input and password_input:
            username_input.send_keys(username)
            password_input.send_keys(password)
            password_input.send_keys(Keys.RETURN)

            # Wait for the page to load and check if the username is present
            try:
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "' + username + '")]')))
                return "yes", ""
            except TimeoutException:
                return "no", "Login failed"
        else:
            return "no", "Username or password input not found"
    else:
        return "no", "Not a login page"

# Add column headers for login status and error
url_sheet.cell(row=1, column=2, value="Login status")
url_sheet.cell(row=1, column=3, value="Error")

# Get the total number of URLs
total_urls = url_sheet.max_row - 1

# Create a tqdm progress bar
pbar = tqdm(total=total_urls, desc="Authentication URLs")

# Create a folder to store cookies
cookies_folder = "user_cookies"
os.makedirs(cookies_folder, exist_ok=True)

# Loop through the URLs and attempt login with a timeout
for row in range(2, total_urls + 2):
    url = url_sheet.cell(row=row, column=1).value
    login_status_cell = url_sheet.cell(row=row, column=2)
    error_cell = url_sheet.cell(row=row, column=3)

    if url:
        for username, password in credentials:
            login_status, error = attempt_login_with_timeout(url, username, password, timeout=60, headers=headers)
            if login_status == "yes":
                login_status_cell.value = "yes"
                error_cell.value = ""

                # Store the cookies in a text file
                cookies_filename = os.path.join(cookies_folder, f"cookies_{username}.txt")
                with open(cookies_filename, 'w') as f:
                    for cookie in driver.get_cookies():
                        f.write(f"{cookie['name']}={cookie['value']}\n")

                break
        else:
            login_status_cell.value = "no"
            error_cell.value = error

    # Update the progress bar
    pbar.update(1)

# Close the progress bar
pbar.close()

# Close the WebDriver
driver.quit()

# Save the modified Excel file
workbook.save("urls_auth.xlsx")
