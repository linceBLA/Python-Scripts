# script to check if URL corresponds to a 200
# This script run through a list of URL to check
# It send a request for each URL and get the response status code
# The status code is register directly in the excel table of the URL list
# This script needs:
# - list of URL in excel table: URL-to-check.xlsx


import openpyxl
import requests

# Load the workbook
workbook = openpyxl.load_workbook("URL-to-check.xlsx")

# Select the appropriate sheet
sheet = workbook["Sheet"]

# Create a function to check the HTTP response status
def check_url(url):
    try:
        response = requests.head(url)
        return response.status_code
    except requests.exceptions.RequestException:
        return None

# Add a new column header for HTTP status
sheet.cell(row=1, column=2, value="HTTP Status")

# Iterate over the rows and update the HTTP status column
for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    url = row[0]
    status = check_url(url)
    sheet.cell(row=idx, column=2).value = status

# Save the modified workbook
workbook.save("URL-to-check.xlsx")
